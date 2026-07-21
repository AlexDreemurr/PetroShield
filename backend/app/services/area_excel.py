from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


DATA_SHEET = "区域数据"
MAX_IMPORT_ROWS = 500
MAX_FILE_SIZE = 5 * 1024 * 1024

HEADERS = [
    "区域名称*", "区域类型*", "风险等级*", "图形类型*", "多边形坐标",
    "圆心X", "圆心Y", "半径(米)", "是否启用*", "负责人", "责任部门",
    "越界告警*", "停留告警*", "停留分钟", "人数告警*", "人数上限", "处置优先级*",
]

AREA_TYPE_VALUES = {"危险": "danger", "危险区": "danger", "danger": "danger", "限制": "restricted", "限制区": "restricted", "restricted": "restricted", "禁入": "prohibited", "禁入区": "prohibited", "prohibited": "prohibited", "普通": "normal", "普通区": "normal", "normal": "normal"}
RISK_LEVEL_VALUES = {"低": "low", "低风险": "low", "low": "low", "中": "medium", "中风险": "medium", "medium": "medium", "高": "high", "高风险": "high", "high": "high"}
SHAPE_VALUES = {"多边形": "polygon", "polygon": "polygon", "圆形": "circle", "circle": "circle"}
PRIORITY_VALUES = {"低": "low", "low": "low", "中": "medium", "medium": "medium", "高": "high", "high": "high", "紧急": "urgent", "urgent": "urgent"}
TRUE_VALUES = {"是", "true", "1", "启用", "开启"}
FALSE_VALUES = {"否", "false", "0", "停用", "关闭"}


class AreaExcelError(ValueError):
    def __init__(self, errors: list[dict[str, Any]]):
        self.errors = errors
        super().__init__("区域表格校验失败")


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _enum(value: Any, values: dict[str, str], field: str) -> str:
    raw = _text(value)
    result = values.get(raw) or values.get(raw.lower())
    if not result:
        raise ValueError(f"{field}取值不正确：{raw or '空'}")
    return result


def _boolean(value: Any, field: str) -> bool:
    raw = _text(value).lower()
    if raw in TRUE_VALUES:
        return True
    if raw in FALSE_VALUES:
        return False
    raise ValueError(f"{field}只能填写是或否")


def _number(value: Any, field: str, *, minimum: float = 0, integer: bool = False) -> float | int:
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field}必须是数字") from exc
    if number < minimum:
        raise ValueError(f"{field}不能小于 {minimum:g}")
    return int(number) if integer else number


def _coordinates(value: Any) -> list[dict[str, float]]:
    raw = _text(value).replace("；", ";").replace("，", ",")
    if not raw:
        return []
    points: list[dict[str, float]] = []
    for index, pair in enumerate(raw.split(";"), start=1):
        values = [part.strip() for part in pair.split(",")]
        if len(values) != 2:
            raise ValueError(f"多边形第 {index} 个坐标应为 x,y")
        try:
            points.append({"x": float(values[0]), "y": float(values[1])})
        except ValueError as exc:
            raise ValueError(f"多边形第 {index} 个坐标不是数字") from exc
    if len(points) < 3:
        raise ValueError("多边形至少需要 3 个坐标点")
    return points


def parse_area_workbook(content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise AreaExcelError([{"row": None, "field": "文件", "message": "无法读取工作簿，请使用平台模板生成的 .xlsx 文件"}]) from exc

    if DATA_SHEET not in workbook.sheetnames:
        raise AreaExcelError([{"row": None, "field": "工作表", "message": f"缺少“{DATA_SHEET}”工作表"}])
    sheet = workbook[DATA_SHEET]
    actual_headers = [_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    missing = [header for header in HEADERS if header not in actual_headers]
    if missing:
        raise AreaExcelError([{"row": 1, "field": "表头", "message": f"缺少列：{'、'.join(missing)}"}])
    indexes = {header: actual_headers.index(header) for header in HEADERS}

    areas: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    names: set[str] = set()
    processed_rows = 0
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(_text(value) for value in row):
            continue
        processed_rows += 1
        if processed_rows > MAX_IMPORT_ROWS:
            errors.append({"row": row_number, "field": "文件", "message": f"单次最多导入 {MAX_IMPORT_ROWS} 个区域"})
            break

        def cell(header: str) -> Any:
            index = indexes[header]
            return row[index] if index < len(row) else None

        try:
            name = _text(cell("区域名称*"))
            if not name:
                raise ValueError("区域名称不能为空")
            if len(name) > 100:
                raise ValueError("区域名称不能超过 100 个字符")
            if name in names:
                raise ValueError("区域名称在表格内重复")
            shape = _enum(cell("图形类型*"), SHAPE_VALUES, "图形类型")
            polygon = _coordinates(cell("多边形坐标")) if shape == "polygon" else []
            center = None
            radius = None
            if shape == "circle":
                center = {
                    "x": _number(cell("圆心X"), "圆心X"),
                    "y": _number(cell("圆心Y"), "圆心Y"),
                }
                radius = _number(cell("半径(米)"), "半径", minimum=0.01)

            dwell_enabled = _boolean(cell("停留告警*"), "停留告警")
            capacity_enabled = _boolean(cell("人数告警*"), "人数告警")
            dwell_minutes = _number(cell("停留分钟"), "停留分钟", minimum=1, integer=True) if dwell_enabled else 30
            max_people = _number(cell("人数上限"), "人数上限", minimum=0, integer=True) if capacity_enabled else 10
            areas.append({
                "name": name,
                "type": _enum(cell("区域类型*"), AREA_TYPE_VALUES, "区域类型"),
                "risk_level": _enum(cell("风险等级*"), RISK_LEVEL_VALUES, "风险等级"),
                "shape": shape,
                "polygon": polygon,
                "center": center,
                "radius": radius,
                "enabled": _boolean(cell("是否启用*"), "是否启用"),
                "manager_name": _text(cell("负责人")) or None,
                "manager_department": _text(cell("责任部门")) or None,
                "rules": {
                    "cross_boundary": _boolean(cell("越界告警*"), "越界告警"),
                    "dwell_enabled": dwell_enabled,
                    "dwell_minutes": dwell_minutes,
                    "capacity_enabled": capacity_enabled,
                    "max_people": max_people,
                    "priority": _enum(cell("处置优先级*"), PRIORITY_VALUES, "处置优先级"),
                },
            })
            names.add(name)
        except ValueError as exc:
            errors.append({"row": row_number, "field": "区域数据", "message": str(exc)})

    workbook.close()
    if not areas and not errors:
        errors.append({"row": None, "field": "区域数据", "message": "工作表中没有可导入的区域"})
    if errors:
        raise AreaExcelError(errors[:50])
    return areas


def build_area_template() -> bytes:
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = DATA_SHEET
    data_sheet.append(HEADERS)
    data_sheet.freeze_panes = "A2"
    data_sheet.auto_filter.ref = f"A1:Q1"
    widths = [20, 13, 13, 13, 36, 11, 11, 13, 13, 15, 18, 13, 13, 13, 13, 13, 15]
    for index, width in enumerate(widths, start=1):
        data_sheet.column_dimensions[data_sheet.cell(1, index).column_letter].width = width
    for cell in data_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    validations = [
        ("B2:B501", '"危险,限制,禁入,普通"'), ("C2:C501", '"低,中,高"'),
        ("D2:D501", '"多边形,圆形"'), ("I2:I501", '"是,否"'),
        ("L2:L501", '"是,否"'), ("M2:M501", '"是,否"'),
        ("O2:O501", '"是,否"'), ("Q2:Q501", '"低,中,高,紧急"'),
    ]
    for cell_range, formula in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        data_sheet.add_data_validation(validation)
        validation.add(cell_range)

    example_sheet = workbook.create_sheet("填写示例")
    example_sheet.append(HEADERS)
    example_sheet.append(["一号罐区", "危险", "高", "多边形", "220,170;300,165;315,235;225,240", "", "", "", "是", "张三", "HSE部", "是", "是", 30, "是", 20, "高"])
    example_sheet.append(["应急集合点", "普通", "低", "圆形", "", 410, 260, 35, "是", "李四", "生产部", "是", "否", "", "否", "", "中"])
    for cell in example_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")

    guide = workbook.create_sheet("填写说明")
    guide_rows = [
        ["项目", "规定"],
        ["上传文件", "仅支持 .xlsx；请在“区域数据”工作表中填写，勿修改带 * 的表头。"],
        ["多边形坐标", "按 x,y;x,y;x,y 格式填写，至少 3 个点，例如 220,170;300,165;315,235。"],
        ["圆形坐标", "多边形坐标留空，填写圆心X、圆心Y和大于 0 的半径。"],
        ["追加模式 a", "保留现有区域并新增表内区域；区域名称不能与现有区域重复。"],
        ["覆盖模式 w", "校验全部通过后清空现有区域并按表格重建；人员和设备解除旧区域关联，历史告警保留。"],
        ["布尔字段", "填写 是/否，也支持 true/false、1/0。"],
        ["数量限制", f"单次最多 {MAX_IMPORT_ROWS} 个区域，文件最大 5 MB。"],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 100
    for cell in guide[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
    for row in guide.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()
